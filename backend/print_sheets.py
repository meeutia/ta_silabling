import openpyxl

def main():
    file_path = "d:/Unand/TA/1. Project TA/backend/docs/Hasil_Testing_Lampiran_SILABLING_Lulus.xlsx"
    wb = openpyxl.load_workbook(file_path)
    print("Sheet names:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        print(f"\n=== {sheet_name} ===")
        sheet = wb[sheet_name]
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > 5:
                break
            print(f"Row {i+1}: {row}")

if __name__ == "__main__":
    main()
