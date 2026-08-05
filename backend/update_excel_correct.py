import openpyxl

def main():
    file_path = "d:/Unand/TA/1. Project TA/backend/docs/Hasil_Testing_Lampiran_SILABLING_Lulus.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    # Update Ringkasan
    if "Ringkasan" in wb.sheetnames:
        ringkasan = wb["Ringkasan"]
        ringkasan.cell(row=5, column=2, value=17) # Unit Suites
        ringkasan.cell(row=5, column=3, value=169) # Unit Cases
        ringkasan.cell(row=5, column=4, value="169 lulus, 0 gagal")
        
        ringkasan.cell(row=6, column=2, value=8) # Integration Suites
        ringkasan.cell(row=6, column=3, value=85) # Integration Cases
        ringkasan.cell(row=6, column=4, value="85 lulus, 0 gagal")
        
    # Update Lampiran H - Unit
    if "Lampiran H - Unit" in wb.sheetnames:
        unit_sheet = wb["Lampiran H - Unit"]
        unit_sheet.cell(row=2, column=1, value="Hasil pengujian: 17 test suite dan 169 test case lulus.")
        
        # Find last row
        last_row = unit_sheet.max_row
        last_no = unit_sheet.cell(row=last_row, column=1).value
        if not isinstance(last_no, int):
            last_no = 20 # fallback
            
        new_unit_rows = [
            ["Layanan Permintaan Subkontrak", "Mengajukan Permintaan Subkontrak", "ID Parameter dan FPM valid", "Permintaan berhasil dibuat (SELESAI/MENUNGGU_ADMIN)", "Permintaan berhasil", "Lulus"],
            [None, "Menyetujui Permintaan Subkontrak", "Payload tarif dan metode valid", "Metode berhasil ditambahkan ke database", "Metode ditambahkan", "Lulus"],
            ["Utilitas Duplikasi", "Deteksi Duplikasi (Aturan Baru)", "Maksud Uji / Baku Mutu berbeda", "Dianggap duplikat (aturan disederhanakan)", "Dianggap duplikat", "Lulus"]
        ]
        
        for row_data in new_unit_rows:
            last_no += 1
            last_row += 1
            unit_sheet.cell(row=last_row, column=1, value=last_no)
            unit_sheet.cell(row=last_row, column=2, value=row_data[0])
            unit_sheet.cell(row=last_row, column=3, value=row_data[1])
            unit_sheet.cell(row=last_row, column=4, value=row_data[2])
            unit_sheet.cell(row=last_row, column=5, value=row_data[3])
            unit_sheet.cell(row=last_row, column=6, value=row_data[4])
            unit_sheet.cell(row=last_row, column=7, value=row_data[5])
            
    # Update Lampiran I - Integration
    if "Lampiran I - Integration" in wb.sheetnames:
        int_sheet = wb["Lampiran I - Integration"]
        int_sheet.cell(row=2, column=1, value="Hasil pengujian: 8 test suite dan 85 test case lulus.")
        
        last_row = int_sheet.max_row
        last_no = int_sheet.cell(row=last_row, column=1).value
        if not isinstance(last_no, int):
            last_no = 15 # fallback
            
        new_int_rows = [
            ["Alur Subkontrak (API)", "Kasi mengajukan subkontrak", "POST /api/kasi/subcontract dengan token Kasi", "Status 201 Created", "Return 201", "Lulus"],
            [None, "Admin menyetujui subkontrak", "POST /api/admin/subcontract/:id/approve", "Status 200 OK dan metode terbuat", "Return 200", "Lulus"],
            ["Revisi Permohonan", "Mencegah Update Permohonan Lama", "PUT /api/customer/requests/:id", "Status 410 Gone", "Return 410", "Lulus"]
        ]
        
        for row_data in new_int_rows:
            last_no += 1
            last_row += 1
            int_sheet.cell(row=last_row, column=1, value=last_no)
            int_sheet.cell(row=last_row, column=2, value=row_data[0])
            int_sheet.cell(row=last_row, column=3, value=row_data[1])
            int_sheet.cell(row=last_row, column=4, value=row_data[2])
            int_sheet.cell(row=last_row, column=5, value=row_data[3])
            int_sheet.cell(row=last_row, column=6, value=row_data[4])
            int_sheet.cell(row=last_row, column=7, value=row_data[5])
            
    wb.save(file_path)
    print("Excel file fully updated with unit and integration stats.")

if __name__ == "__main__":
    main()
