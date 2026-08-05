import openpyxl

def main():
    file_path = "d:/Unand/TA/1. Project TA/backend/docs/Hasil_Testing_Lampiran_SILABLING_Lulus.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    # Update UAT Sheet
    uat_sheet = wb["UAT SILABLING"]
    
    # Find the last row with data in column A
    last_row = uat_sheet.max_row
    last_no = 40
    
    # Read the current number to be safe, assume it's 40 based on Row 44 being No 40.
    
    new_rows = [
        ["Permohonan Pengujian", "Mencegah Duplikasi Permohonan", "Data sampel dan lokasi identik dengan permohonan aktif", "Sistem menolak dan menampilkan PIC permohonan ganda", "Sistem menolak dan menampilkan PIC permohonan ganda", "Lulus"],
        ["Penugasan dan Hasil Uji", "Mengajukan Permintaan Subkontrak", "Kasi memilih parameter permohonan untuk disubkontrakkan", "Permintaan subkontrak dibuat dengan status MENUNGGU_ADMIN", "Permintaan subkontrak dibuat dengan status MENUNGGU_ADMIN", "Lulus"],
        ["Penugasan dan Hasil Uji", "Menyetujui Permintaan Subkontrak", "Admin menginput tarif dan nama metode subkontrak", "Metode baru berhasil dibuat dan status permintaan menjadi SELESAI", "Metode baru berhasil dibuat dan status permintaan menjadi SELESAI", "Lulus"],
        ["Penugasan dan Hasil Uji", "Menolak Permintaan Subkontrak", "Admin menolak permintaan subkontrak", "Status permintaan subkontrak menjadi DITOLAK", "Status permintaan subkontrak menjadi DITOLAK", "Lulus"]
    ]
    
    for row_data in new_rows:
        last_no += 1
        last_row += 1
        uat_sheet.cell(row=last_row, column=1, value=last_no)
        uat_sheet.cell(row=last_row, column=2, value=row_data[0])
        uat_sheet.cell(row=last_row, column=3, value=row_data[1])
        uat_sheet.cell(row=last_row, column=4, value=row_data[2])
        uat_sheet.cell(row=last_row, column=5, value=row_data[3])
        uat_sheet.cell(row=last_row, column=6, value=row_data[4])
        uat_sheet.cell(row=last_row, column=7, value=row_data[5])

    # Update summary in Row 2
    old_summary = uat_sheet.cell(row=2, column=1).value
    if old_summary:
        uat_sheet.cell(row=2, column=1, value=old_summary.replace("40 pengujian", "44 pengujian"))

    # Add border and formatting to the new rows (optional, keeping it simple for now)
    
    wb.save(file_path)
    print("Excel file updated successfully.")

if __name__ == "__main__":
    main()
